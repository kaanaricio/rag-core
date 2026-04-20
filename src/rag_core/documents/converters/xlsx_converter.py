"""XLSX converter with both formula and computed value modes.

Improvements over AirWeave:
- Supports BOTH data_only=True (computed values) and data_only=False (formulas)
- Handles merged cells gracefully
- Multiple sheets with clear separation
- Configurable row/column limits for large files
"""

from __future__ import annotations

import asyncio
import io
import logging
from rag_core.config.env_access import get_env as config_get_env
from itertools import zip_longest
from typing import Any, Dict, List, Tuple

from .base import BaseConverter, ConversionResult, render_markdown_table, score_text_quality

logger = logging.getLogger(__name__)


def _chart_title(chart: Any) -> str:
    """Best-effort extraction of an openpyxl chart title."""
    title = getattr(chart, "title", None)
    if title is None:
        return ""
    # OpenPyXL can represent title as rich text objects; string conversion
    # is robust enough for indexing/metadata.
    text = str(title).strip()
    if text == "None":
        return ""
    return text


class XlsxConverter(BaseConverter):
    """Converts XLSX files to markdown tables.

    Unlike AirWeave which only uses data_only=False, we support both modes
    and include computed values when available.
    """

    format_name = "xlsx"

    def __init__(
        self,
        *,
        max_rows: int = 0,
        max_cols: int = 0,
        include_formulas: bool = False,
    ) -> None:
        self._rows_per_chunk = max_rows or int(config_get_env("LOCAL_PARSE_XLSX_MAX_ROWS", "500"))
        self._max_total_rows = int(config_get_env("LOCAL_PARSE_XLSX_MAX_TOTAL_ROWS", "5000"))
        self._max_cols = max_cols or int(config_get_env("LOCAL_PARSE_XLSX_MAX_COLS", "50"))
        self._include_formulas = include_formulas

    async def convert(
        self,
        file_bytes: bytes,
        filename: str,
        mime_type: str,
    ) -> ConversionResult:
        """Convert XLSX to markdown tables."""
        from openpyxl import load_workbook  # type: ignore[import-not-found]

        def _extract() -> ConversionResult:
            sections: List[str] = []
            sheet_count = 0
            figure_items: List[Dict[str, Any]] = []
            chart_map: Dict[str, List[Any]] = {}
            workbook_rows_consumed = 0
            workbook_truncated = False

            # Load with computed values first (most useful for downstream)
            try:
                wb_data = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
            except Exception as exc:
                logger.warning("Failed to open XLSX %s: %s", filename, exc)
                return ConversionResult(
                    metadata={"parser": "local:openpyxl", "error": str(exc)},
                )

            wb_formula = None
            try:
                wb_formula = load_workbook(io.BytesIO(file_bytes), data_only=False, read_only=True)
            except Exception:
                logger.debug("Could not load formula workbook for %s", filename)

            # Read chart metadata in non-read-only mode; openpyxl does not expose
            # chart objects on read-only worksheets.
            wb_chart = None
            try:
                wb_chart = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
                for chart_sheet in wb_chart.worksheets:
                    chart_map[chart_sheet.title] = list(getattr(chart_sheet, "_charts", []))
            except Exception:
                logger.debug("Could not load chart metadata for %s", filename)

            for sheet_index, sheet in enumerate(wb_data.worksheets):
                sheet_count += 1
                formula_sheet = None
                if wb_formula is not None:
                    try:
                        formula_sheet = wb_formula[sheet.title]
                    except Exception:
                        # Formula workbooks can diverge from the data workbook; keep extracting cell values only.
                        formula_sheet = None

                remaining_row_budget = max(self._max_total_rows - workbook_rows_consumed, 0)
                rows_data, rows_consumed, truncated = self._extract_sheet(
                    sheet,
                    formula_sheet=formula_sheet,
                    remaining_total_rows=remaining_row_budget,
                )
                workbook_rows_consumed += rows_consumed
                workbook_truncated = workbook_truncated or truncated
                charts = chart_map.get(sheet.title, [])
                if not rows_data and not charts:
                    if truncated:
                        sections.append(
                            "\n\n".join(
                                [
                                    "## Sheet: %s" % sheet.title,
                                    "*[truncated after workbook limit of %d rows]*"
                                    % self._max_total_rows,
                                ]
                            )
                        )
                    continue

                chunk_size = max(1, self._rows_per_chunk)
                row_groups: List[Tuple[int, int, List[List[str]]]] = []
                for start_idx in range(0, len(rows_data), chunk_size):
                    chunk_rows = rows_data[start_idx : start_idx + chunk_size]
                    if not chunk_rows:
                        continue
                    start_row = start_idx + 1
                    end_row = start_idx + len(chunk_rows)
                    row_groups.append((start_row, end_row, chunk_rows))

                for group_index, (start_row, end_row, group_rows) in enumerate(row_groups):
                    if len(row_groups) > 1:
                        heading = "## Sheet: %s (Rows %d-%d)" % (
                            sheet.title,
                            start_row,
                            end_row,
                        )
                    else:
                        heading = "## Sheet: %s" % sheet.title

                    section_parts: List[str] = [heading]
                    if len(group_rows) == 1:
                        for cell in group_rows[0]:
                            if cell.strip():
                                section_parts.append("- %s" % cell.strip())
                    else:
                        section_parts.append(render_markdown_table(group_rows))

                    if truncated and group_index == len(row_groups) - 1:
                        section_parts.append(
                            "*[truncated after workbook limit of %d rows]*"
                            % self._max_total_rows
                        )

                    sections.append("\n\n".join(section_parts))

                if charts:
                    chart_parts: List[str] = ["## Sheet: %s Charts" % sheet.title, "### Charts"]
                    for chart_index, chart in enumerate(charts):
                        chart_title = _chart_title(chart) or (
                            "Chart %d on sheet %s" % (chart_index + 1, sheet.title)
                        )
                        chart_parts.append("- %s" % chart_title)
                        figure_items.append(
                            {
                                "figure_id": "fig:sheet:%d:chart:%d"
                                % (sheet_index + 1, chart_index + 1),
                                "page_index": sheet_index,
                                "label": "Sheet %s Chart %d" % (sheet.title, chart_index + 1),
                                "description": chart_title,
                                "metadata": {
                                    "source": "xlsx:chart",
                                    "sheet": sheet.title,
                                    "sheet_name": sheet.title,
                                },
                            }
                        )
                    sections.append("\n\n".join(chart_parts))

            if wb_formula:
                try:
                    wb_formula.close()
                except Exception:
                    # Workbook cleanup is best-effort; keep parsed content even if close fails.
                    pass
            if wb_chart:
                try:
                    wb_chart.close()
                except Exception:
                    # Workbook cleanup is best-effort; keep parsed content even if close fails.
                    pass
            try:
                wb_data.close()
            except Exception:
                # Workbook cleanup is best-effort; keep parsed content even if close fails.
                pass

            content = "\n\n".join(sections)
            quality = score_text_quality(content)

            metadata: Dict[str, Any] = {
                "parser": "local:openpyxl",
                "sheet_count": sheet_count,
                "needs_ocr": False,
            }
            if workbook_truncated:
                metadata["row_truncated"] = True
                metadata["row_limit"] = self._max_total_rows
                metadata["row_limit_scope"] = "workbook"
                metadata["rows_emitted"] = workbook_rows_consumed
            if figure_items:
                metadata["figure_items"] = figure_items
                metadata["figure_count"] = len(figure_items)

            return ConversionResult(
                content=content,
                metadata=metadata,
                quality=quality,
            )

        return await asyncio.to_thread(_extract)

    def _extract_sheet(
        self,
        sheet: Any,
        *,
        formula_sheet: Any | None = None,
        remaining_total_rows: int | None = None,
    ) -> tuple[List[List[str]], int, bool]:
        """Extract rows from a single sheet with workbook-scoped row limits."""
        all_rows: List[List[str]] = []
        row_count = 0
        truncated = False
        max_rows_for_sheet = self._max_total_rows if remaining_total_rows is None else max(
            remaining_total_rows, 0
        )

        formula_iter = (
            formula_sheet.iter_rows(values_only=True) if formula_sheet is not None else []
        )
        for row, formula_row in zip_longest(
            sheet.iter_rows(values_only=True),
            formula_iter,
            fillvalue=(),
        ):
            row_count += 1
            if row_count > max_rows_for_sheet:
                truncated = True
                break

            cells: List[str] = []
            values = tuple(row) if isinstance(row, tuple) else tuple(row or ())
            formula_values = (
                tuple(formula_row) if isinstance(formula_row, tuple) else tuple(formula_row or ())
            )
            for j in range(max(len(values), len(formula_values))):
                if j >= self._max_cols:
                    cells.append("[...]")
                    break
                cells.append(
                    self._render_cell(
                        value=(values[j] if j < len(values) else None),
                        formula_value=(formula_values[j] if j < len(formula_values) else None),
                    )
                )

            # Skip entirely empty rows
            if any(c.strip() for c in cells):
                all_rows.append(cells)

        return all_rows, min(row_count, max_rows_for_sheet), truncated

    def _render_cell(self, *, value: Any, formula_value: Any) -> str:
        formula = (
            formula_value
            if isinstance(formula_value, str) and formula_value.startswith("=")
            else None
        )
        if value is None and formula:
            return formula
        if self._include_formulas and formula:
            if value is None:
                return formula
            return f"{value} [formula: {formula}]"
        return "" if value is None else str(value)
